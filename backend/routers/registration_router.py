"""
Public candidate self-registration + bulk Excel upload.
"""
import io
import secrets
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel, EmailStr
from typing import Optional, List

from database import get_db
import models
from auth import hash_password, generate_roll_number, require_permission

router = APIRouter(prefix="/register", tags=["registration"])


class SelfRegisterRequest(BaseModel):
    full_name: str
    email: EmailStr
    password: str
    phone: Optional[str] = None
    aadhaar: Optional[str] = None
    customer_code: Optional[str] = None   # college code — maps to college if provided


class SelfRegisterResponse(BaseModel):
    message: str
    cid: str
    email: str


@router.post("/candidate", response_model=SelfRegisterResponse)
def self_register(payload: SelfRegisterRequest, db: Session = Depends(get_db)):
    """
    Public self-registration for candidates.
    - If customer_code is provided, maps to that college.
    - Account starts as pending approval (is_approved=False).
    - Duplicate email, phone, Aadhaar are blocked.
    """
    # Duplicate email check
    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")

    # Duplicate phone check
    if payload.phone:
        if db.query(models.User).filter(models.User.phone == payload.phone).first():
            raise HTTPException(status_code=400, detail="Phone number already registered")

    # Duplicate Aadhaar check
    if payload.aadhaar:
        if db.query(models.User).filter(models.User.aadhaar == payload.aadhaar).first():
            raise HTTPException(status_code=400, detail="Aadhaar number already registered")

    # Resolve college from customer code
    college_id = None
    if payload.customer_code:
        college = db.query(models.College).filter(
            models.College.code == payload.customer_code.upper(),
            models.College.is_active == True,
        ).first()
        if not college:
            raise HTTPException(status_code=404, detail="Invalid customer code")
        # Check seat quota
        current_count = db.query(models.User).filter(
            models.User.college_id == college.id,
            models.User.role == models.UserRole.student,
            models.User.is_deleted == False,
        ).count()
        if current_count >= college.seat_limit:
            raise HTTPException(status_code=400, detail="Institution seat limit reached")
        college_id = college.id

    cid = generate_roll_number(db)

    user = models.User(
        email=payload.email,
        hashed_password=hash_password(payload.password),
        role=models.UserRole.student,
        roll_number=cid,
        full_name=payload.full_name,
        phone=payload.phone,
        aadhaar=payload.aadhaar,
        college_id=college_id,
        is_approved=False,       # requires admin approval
        is_active=False,
        is_email_verified=False, # email verification pending
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    return SelfRegisterResponse(
        message="Registration successful. Your account is pending admin approval.",
        cid=cid,
        email=payload.email,
    )


@router.post("/bulk-upload")
async def bulk_upload_candidates(
    file: UploadFile = File(...),
    college_id: Optional[int] = None,
    default_password: str = "Candidate@123",
    db: Session = Depends(get_db),
    _: models.User = Depends(require_permission("manage_students")),
):
    """
    Bulk upload candidates from Excel (.xlsx).
    Expected columns: full_name, email, phone, aadhaar, course_code, batch_name
    Returns summary of created/skipped rows.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    # Read header row
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    required = {'full_name', 'email'}
    missing = required - set(headers)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {missing}. Required: full_name, email"
        )

    def col(row, name):
        try:
            idx = headers.index(name)
            val = row[idx].value
            return str(val).strip() if val is not None else None
        except (ValueError, IndexError):
            return None

    created = []
    skipped = []

    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue

        email = col(row, 'email')
        full_name = col(row, 'full_name')
        phone = col(row, 'phone')
        aadhaar = col(row, 'aadhaar')
        course_code = col(row, 'course_code')
        batch_name = col(row, 'batch_name')

        if not email or not full_name:
            skipped.append({"row": email or "?", "reason": "Missing email or full_name"})
            continue

        # Duplicate checks
        if db.query(models.User).filter(models.User.email == email).first():
            skipped.append({"row": email, "reason": "Email already exists"})
            continue
        if phone and db.query(models.User).filter(models.User.phone == phone).first():
            skipped.append({"row": email, "reason": "Phone already exists"})
            continue
        if aadhaar and db.query(models.User).filter(models.User.aadhaar == aadhaar).first():
            skipped.append({"row": email, "reason": "Aadhaar already exists"})
            continue

        # Resolve course
        course_id = None
        if course_code:
            course = db.query(models.Course).filter(
                models.Course.code == course_code.upper()
            ).first()
            if course:
                course_id = course.id

        cid = generate_roll_number(db)
        user = models.User(
            email=email,
            hashed_password=hash_password(default_password),
            role=models.UserRole.student,
            roll_number=cid,
            full_name=full_name,
            phone=phone,
            aadhaar=aadhaar,
            college_id=college_id,
            course_id=course_id,
            is_approved=True,    # bulk upload by admin = auto-approved
            is_active=True,
            is_email_verified=True,
        )
        db.add(user)
        db.flush()

        # Auto-assign to batch if batch_name provided
        if batch_name:
            batch = db.query(models.Batch).filter(
                models.Batch.name == batch_name,
                models.Batch.is_active == True,
            ).first()
            if batch:
                existing = db.query(models.BatchMember).filter(
                    models.BatchMember.batch_id == batch.id,
                    models.BatchMember.student_id == user.id,
                ).first()
                if not existing:
                    db.add(models.BatchMember(batch_id=batch.id, student_id=user.id))

        created.append({"email": email, "cid": cid, "full_name": full_name})

    db.commit()

    return {
        "created": len(created),
        "skipped": len(skipped),
        "message": f"{len(created)} candidates created, {len(skipped)} skipped.",
        "candidates": created,
        "skipped_details": skipped,
    }
